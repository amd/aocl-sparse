"""Copyright (C) 2020, Advanced Micro Devices, Inc. All Rights Reserved"""
import re
import subprocess
import random
import os
from datetime import datetime
import shutil
from collections import OrderedDict
import platform
import openpyxl
import yaml
from openpyxl.styles import Side


class SparseExecution:

    @staticmethod
    def read_sparse_commands():
        """
        :Method Name: read_sparse_commands
        :Description: reads the commands from the input file
        :parameter  : None
        :return     : Command"""

        try:
            # book = (SparseExecution.yaml_inputs())[0]
            sheet = (SparseExecution.yaml_inputs())[1]
            rows = sheet.max_row
            bench_path_linux = (SparseExecution.yaml_inputs())[5]
            print("bench_path_linux", bench_path_linux)
            for i in range(2, rows + 1):
                command = sheet.cell(row=i, column=12).value
                expected = sheet.cell(row=i, column=13).value
                if plt == "Windows":
                    if command:
                        SparseExecution.add_commandline_arguments(command[2:], i, expected)
                    else:
                        print("Execution completed")
                        break
                elif plt == "Linux":
                    if command:
                        bench_str = "./aoclsparse-bench" #added changes for running script from tests/scripts in linux
                        command = command.replace(bench_str, bench_path_linux+"aoclsparse-bench")
                        SparseExecution.add_commandline_arguments(command, i, expected)
                    else:
                        print("Execution completed")
                        break

        except FileNotFoundError:
            print("Please keep inputs.yaml and aocl-sparse_test_plan.xlsx in the aoclsparse-bench.exe directory")

    @staticmethod
    def add_commandline_arguments(command, i, expected):
        """
        :Method Name: add_commandline_arguments
        :Description: adds value to the m, n, z, alpha and beta arguments
                      according to the inputs provided in the input file
        :parameter  : command, i, expected
        :return     : command"""

        sheet = (SparseExecution.yaml_inputs())[1]
        command = str(command)
        z_value = sheet.cell(row=i, column=8).value
        if z_value == "random":
            z_value = random.randrange(100, 1000, 100)
        elif z_value == ">100":
            z_value = random.randrange(101, 500, 10)

        matrix = sheet.cell(row=i, column=10).value
        m_str, n_str, alpha_str, beta_str, z_str, dim_str = "--sizem=<>", "--sizen=<>", "--alpha=<>", "--beta=<>", "--sizennz=<>", "--blockdim=<>"
        mtx_str = "--mtx=<>"

        if matrix == "random":
            value = random.randrange(100, 1001, 100)
            for raw_value in ((m_str, "--sizem=" + str(value)), (n_str, "--sizen=" + str(value)),
                      (alpha_str, "--alpha=" + str(1)), (beta_str, "--beta=" + str(0))):
                command = command.replace(*raw_value)
                z_calculation = (value * value * z_value) // 100
                command = command.replace(z_str, "--sizennz=" + str(z_calculation))
                z_str = "--sizennz=" + str(z_calculation)

            SparseExecution.command_execution(command, expected)

        elif matrix == 0:
            for raw_value in ((m_str, "--sizem=" + str(0)), (n_str, "--sizen=" + str(0)), (alpha_str, "--alpha=" + str(1)),
                      (beta_str, "--beta=" + str(0))):
                command = command.replace(*raw_value)
                command = command.replace(z_str, "--sizennz=" + str(100))

            SparseExecution.command_execution(command, expected)

        elif matrix == "NA":
            # exec_dir = os.getcwd()
            if plt == "Windows":
                cmd_path = '../../../../'
            elif  plt == "Linux":
                cmd_path = '../../'
            # os.chdir(exec_dir)
            mtx_path = os.path.join(cmd_path, SparseExecution.yaml_inputs()[4] )
            files = [f for f in os.listdir(mtx_path) if re.search('.mtx', f)]
            if len(files) != 0:
                if plt == "Windows":
                    for file_iter in range(len(files)):
                        command = (SparseExecution.yaml_inputs())[2] + " " + sheet.cell(row=i, column=12).value
                        command = command.replace(mtx_str, "--mtx=" + mtx_path + files[file_iter])
                        dim_str = "--blockdim=<>"
                        if dim_str in command: # Command contains --blockdim=<>
                            dim_value = sheet.cell(row=i, column=14).value
                            dim_range = dim_value.split("/")
                            for dim_loop in dim_range:
                                command = command.replace(dim_str, "--blockdim=" + str(dim_loop))
                                dim_str = "--blockdim=" + str(dim_loop)
                                SparseExecution.command_execution(command, expected)
                        else:
                            SparseExecution.command_execution(command, expected)
                elif plt == "Linux":
                    for file_iter in range(len(files)):
                        command = (SparseExecution.yaml_inputs())[3] + " " + sheet.cell(row=i, column=12).value
                        bench_path_linux = (SparseExecution.yaml_inputs())[5]
                        bench_str = "./aoclsparse-bench" #added changes for running script from tests/scripts in linux
                        command = command.replace(bench_str, bench_path_linux+"aoclsparse-bench")
                        command = command.replace(mtx_str, "--mtx=" + mtx_path + files[file_iter])
                        dim_str = "--blockdim=<>" #change done for failing  bsrmv std matrices 
                        if dim_str in command: # Command contains --blockdim=<>
                            dim_value = sheet.cell(row=i, column=14).value
                            dim_range = dim_value.split("/")
                            for dim_loop in dim_range:
                                command = command.replace(dim_str, "--blockdim=" + str(dim_loop))
                                dim_str = "--blockdim=" + str(dim_loop)
                                SparseExecution.command_execution(command, expected)
                        else:
                            SparseExecution.command_execution(command, expected)
            else:
                print("No .mtx files are found in ", mtx_path, " directory")

        elif matrix:
            matrix_range = matrix.split("to")

            for j in range(int(matrix_range[0]), int(matrix_range[1]) + 1, 100):
                # this loop runs for 10 iterations in multiples of 100
                # and adds value to the m, n arguments
                for n in range(int(matrix_range[0]), int(matrix_range[1]) + 1, 100):
                    command = command.replace(m_str, "--sizem=" + str(j))
                    command = command.replace(n_str, "--sizen=" + str(n))
                    m_str, n_str = "--sizem=" + str(j), "--sizen=" + str(n)
                    # for calculating z value as percentage of m * n
                    z_calculation = (j * n * z_value) // 100
                    command = command.replace(z_str, "--sizennz=" + str(z_calculation))
                    z_str = "--sizennz=" + str(z_calculation)

                    for k in range(-2, 3):
                        # this loop runs for 4 iterations from -2 to 2 and adds value to alpha and beta arguments
                        command = command.replace(alpha_str, "--alpha=" + str(k))
                        command = command.replace(beta_str, "--beta=" + str(k))
                        alpha_str, beta_str = "--alpha=" + str(k), "--beta=" + str(k)

                        if dim_str in command: # Command contains --blockdim=<>
                            dim_value = sheet.cell(row=i, column=14).value
                            dim_range = dim_value.split("/")
                            for dim_loop in dim_range:
                                command = command.replace(dim_str, "--blockdim=" + str(dim_loop))
                                dim_str = "--blockdim=" + str(dim_loop)
                                SparseExecution.command_execution(command, expected)
                        else:
                            SparseExecution.command_execution(command, expected)


    @staticmethod
    def command_execution(command, expected):
        """
        :Method Name: command_execution
        :Description: output.xlsx is created to add command output and command is executed
        :parameter  : command, expected
        :return     : None"""

        book = openpyxl.load_workbook('output.xlsx')
        sheet = book.active
        rows = sheet.max_row
        print(command)
        rows += 1
        raw_output = subprocess.getoutput(command)
        print(raw_output, '\n')
        out = raw_output.split()

        if len(out) == 20:  # valid output
            AddSparseResults.add_output_to_xl(rows, out, sheet, command, expected, book)

        elif "error" in str(out):  # error occurred
            sheet.cell(row=rows, column=1).value = command
            sheet.cell(row=rows, column=12).value = str(out)
            Validation.add_validation_result(sheet, expected, rows, book)

        elif "ASSERT" in str(out):  # error occurred
            sheet.cell(row=rows, column=1).value = command
            sheet.cell(row=rows, column=12).value = str(out)
            Validation.add_validation_result(sheet, expected, rows, book)

        elif ".mtx" in str(out):
            AddSparseResults.add_output_to_xl(rows, out[5:], sheet, command, expected, book)

        else:
            print("Invalid command\n {}\nPLEASE ENTER A VALID COMMAND\n".format(command))
            sheet.cell(row=rows, column=1).value = command
            sheet.cell(row=rows, column=12).value = 'Invalid instruction'
            Validation.add_validation_result(sheet, expected, rows, book)

        book.save("output.xlsx")

    @staticmethod
    def yaml_inputs():
        """
        :Method Name: yaml_inputs
        :Description: Loads the inputs.yaml file and reads the input variables
        :parameter  : inputs.yaml
        :return     : book, sheet, numa_win, numa_linux"""

        with open(r'inputs.yaml') as file:
            # The FullLoader parameter handles the conversion from YAML
            # scalar values to Python the dictionary format
            input_file = yaml.load(file, Loader=yaml.FullLoader) \

            book = openpyxl.load_workbook(input_file['file_name'])
            sheet = book[input_file['sheet_name']]
            numa_win = input_file['numa_cmd_win']
            numa_linux = input_file['numa_cmd_linux']
            mtx_file_path = input_file['mtx_path']
            bench_path_linux = input_file['bench_path_linux']

            return book, sheet, numa_win, numa_linux, mtx_file_path, bench_path_linux


class AddSparseResults:

    @staticmethod
    def create_xl():
        """
        :Method Name: create_xl
        :Description: creates an output excel file and adds required headers
        :parameter  : None
        :return     : None"""

        book = openpyxl.Workbook()
        sheet = book.active
        sheet['A1'], sheet['B1'], sheet['C1'], sheet['D1'], sheet['E1'], sheet[
            'F1'], sheet['G1'], sheet['H1'], sheet['I1'], sheet['J1'], sheet['K1'], sheet[
            'L1'], sheet[
            'M1'] = "Commands", "M", "N", "nnz", "alpha", "beta", "Gflop/s", "GB/s", "msec ", "iter", "verified", "Error ", "Result "
        header = openpyxl.styles.NamedStyle(name="header")
        header.font = openpyxl.styles.Font(bold=True, size=14)
        header.border = openpyxl.styles.Border(bottom=Side(border_style="thin"))
        header.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
        fill = openpyxl.styles.PatternFill("solid", fgColor="003366FF")
        header_row = sheet[1]
        for cell in header_row:
            cell.style, cell.fill = header, fill

        book.save("output.xlsx")

    @staticmethod
    def add_output_to_xl(i, out, sheet, command, expected, book):
        """
        :Method Name: add_output_to_xl
        :Description: converts the raw output into an ordered dictionary and adds the output to output file
        :parameter  : i, out, sheet, command, expected, book
        :return     : None"""

        middle_index = len(out) // 2
        first_half, second_half = out[:middle_index], out[middle_index:]
        out_dict = OrderedDict((first_half[i], second_half[i]) for i in range(len(first_half)))
        sheet.cell(row=i, column=1).value = command
        sheet.cell(row=i, column=2).value = int(out_dict['M'])
        sheet.cell(row=i, column=3).value = int(out_dict['N'])
        sheet.cell(row=i, column=4).value = int(out_dict['nnz'])
        sheet.cell(row=i, column=5).value = int(float(out_dict['alpha']))
        sheet.cell(row=i, column=6).value = int(float(out_dict['beta']))

        if out_dict['GFlop/s'] == "inf":
            sheet.cell(row=i, column=7).value = out_dict['GFlop/s']
        else:
            sheet.cell(row=i, column=7).value = float(out_dict['GFlop/s'])

        if out_dict['GB/s'] == "inf":
            sheet.cell(row=i, column=8).value = out_dict['GB/s']
        else:
            sheet.cell(row=i, column=8).value = float(out_dict['GB/s'])

        sheet.cell(row=i, column=9).value = float(out_dict['msec'])
        sheet.cell(row=i, column=10).value = int(out_dict['iter'])
        sheet.cell(row=i, column=11).value = out_dict['verified']

        Validation.add_validation_result(sheet, expected, i, book)


class Validation:
    """checks the output and gives either pass or fail result"""

    @staticmethod
    def add_validation_result(sheet, expected, i, book):
        """
        :Method Name: add_validation_result
        :Description: compares the expected output with the actual output
                      if it matches, it appends Pass into the Result column otherwise appends Fail
        :parameter  : sheet, expected, i, book
        :return     : None"""

        if expected == "verified = yes":
            verify_value = sheet.cell(row=i, column=11).value
            if verify_value == "yes":
                sheet.cell(row=i, column=13).value = "Pass"
            else:
                sheet.cell(row=i, column=13).value = "Fail"

        elif expected == "aoclsparse_status_not_implemented":
            verify_value = sheet.cell(row=i, column=12).value
            if verify_value:
                if expected in verify_value:
                    sheet.cell(row=i, column=13).value = "Pass"
            else:
                sheet.cell(row=i, column=13).value = "Fail"

        elif expected == "aoclsparse_status_invalid_size":
            verify_value = sheet.cell(row=i, column=12).value
            if verify_value:
                if expected in verify_value:
                    sheet.cell(row=i, column=13).value = "Pass"
            else:
                sheet.cell(row=i, column=13).value = "Fail"

        elif expected == "verified = yes for all matrices.mtx":
            verify_value = sheet.cell(row=i, column=11).value
            if verify_value == "yes":
                sheet.cell(row=i, column=13).value = "Pass"
            else:
                sheet.cell(row=i, column=13).value = "Fail"

        book.save("output.xlsx")
        book.close()


class RenameOutputFile:
    """Renames output.xlsx with current timestamp"""

    @staticmethod
    def rename_with_datetime():
        """
        :Method Name: add_validation_result
        :Description: This creates a directory in order to store all the execution reports with current timestamp
                      as filename.the execution report is copied to sparse-execution-reports directory and
                      filename is changed to current timestamp
        :parameter  : sheet, expected, i, book
        :return     : None"""

        if not os.path.exists("sparse-execution-reports"):
            os.mkdir("sparse-execution-reports")

        src_dir = os.getcwd()  # get the current working dir
        dest_dir = src_dir + "/sparse-execution-reports"
        current_time = str(datetime.utcnow())
        current_time = "_".join(current_time.split()).replace(":", "-")
        current_time = current_time[:-7]
        os.rename('output.xlsx', str(current_time) + '.xlsx')
        src_file = os.path.join(src_dir, str(current_time) + '.xlsx')
        shutil.move(src_file, dest_dir)  # copy the file to destination dir


if __name__ == "__main__":
    start_time = datetime.now()
    parent_dir = os.getcwd()
    plt = platform.system()
    AddSparseResults.create_xl()
    sparse_object = SparseExecution()
    sparse_object.read_sparse_commands()
    RenameOutputFile.rename_with_datetime()
    end_time = datetime.now()
    total_time = end_time - start_time
    # print("total execution time: ",total_time)
